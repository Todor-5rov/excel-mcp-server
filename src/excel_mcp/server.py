import logging
import os
import shutil
from typing import Any, List, Dict, Optional

from fastmcp import FastMCP
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import uvicorn

# Import exceptions
from excel_mcp.exceptions import (
    ValidationError,
    WorkbookError,
    SheetError,
    DataError,
    FormattingError,
    CalculationError,
    PivotError,
    ChartError
)

# Import from excel_mcp package with consistent _impl suffixes
from excel_mcp.validation import (
    validate_formula_in_cell_operation as validate_formula_impl,
    validate_range_in_sheet_operation as validate_range_impl
)
from excel_mcp.chart import create_chart_in_sheet as create_chart_impl
from excel_mcp.workbook import get_workbook_info
from excel_mcp.data import write_data
from excel_mcp.pivot import create_pivot_table as create_pivot_table_impl
from excel_mcp.tables import create_excel_table as create_table_impl
from excel_mcp.sheet import (
    copy_sheet,
    delete_sheet,
    rename_sheet,
    merge_range,
    unmerge_range,
    get_merged_ranges,
)

# Get project root directory path for log file path.
# When using the stdio transmission method,
# relative paths may cause log files to fail to create
# due to the client's running location and permission issues,
# resulting in the program not being able to run.
# Thus using os.path.join(ROOT_DIR, "excel-mcp.log") instead.

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LOG_FILE = os.path.join(ROOT_DIR, "excel-mcp.log")

# Initialize EXCEL_FILES_PATH variable without assigning a value
EXCEL_FILES_PATH = None

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        # Referring to https://github.com/modelcontextprotocol/python-sdk/issues/409#issuecomment-2816831318
        # The stdio mode server MUST NOT write anything to its stdout that is not a valid MCP message.
        logging.FileHandler(LOG_FILE)
    ],
)
logger = logging.getLogger("excel-mcp")
# Initialize FastMCP server
mcp = FastMCP("excel-mcp", stateless_http=True)

def get_excel_path(filename: str) -> str:
    """Get full path to Excel file.
    
    Args:
        filename: Name of Excel file
        
    Returns:
        Full path to Excel file
        
    Raises:
        ValueError: If path validation fails
    """
    # If filename is already an absolute path, return it
    if os.path.isabs(filename):
        if os.path.exists(filename):
            return filename
        else:
            raise ValueError(f"File not found: {filename}")

    # Check if in SSE mode (EXCEL_FILES_PATH is not None)
    if EXCEL_FILES_PATH is None:
        # Must use absolute path
        raise ValueError(f"Invalid filename: {filename}, must be an absolute path when not in SSE mode")

    # In SSE mode, if it's a relative path, resolve it based on EXCEL_FILES_PATH
    full_path = os.path.join(EXCEL_FILES_PATH, filename)
    if not os.path.exists(full_path):
        raise ValueError(f"File not found: {full_path}")
    
    return full_path

@mcp.tool()
def apply_formula(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """
    Apply Excel formula to cell.
    Excel formula will write to cell with verification.
    """
    try:
        full_path = get_excel_path(filepath)
        # First validate the formula
        validation = validate_formula_impl(full_path, sheet_name, cell, formula)
        if isinstance(validation, dict) and "error" in validation:
            return f"Error: {validation['error']}"
            
        # If valid, apply the formula
        from excel_mcp.calculations import apply_formula as apply_formula_impl
        result = apply_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except ValueError as e:
        return f"File path error: {str(e)}"
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error applying formula: {e}")
        return f"Error applying formula: {str(e)}"

@mcp.tool()
def validate_formula_syntax(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """Validate Excel formula syntax without applying it."""
    try:
        full_path = get_excel_path(filepath)
        result = validate_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating formula: {e}")
        return f"Error validating formula: {str(e)}"

@mcp.tool()
def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None
) -> str:
    """Apply formatting to a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.formatting import format_range as format_range_func
        
        # Convert None values to appropriate defaults for the underlying function
        format_range_func(
            filepath=full_path,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,  # This can be None
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,  # This can be None
            font_color=font_color,  # This can be None
            bg_color=bg_color,  # This can be None
            border_style=border_style,  # This can be None
            border_color=border_color,  # This can be None
            number_format=number_format,  # This can be None
            alignment=alignment,  # This can be None
            wrap_text=wrap_text,
            merge_cells=merge_cells,
            protection=protection,  # This can be None
            conditional_format=conditional_format  # This can be None
        )
        return "Range formatted successfully"
    except (ValidationError, FormattingError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error formatting range: {e}")
        return f"Error formatting range: {str(e)}"

@mcp.tool()
def read_data_from_excel(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False
) -> str:
    """
    Read data from Excel worksheet with cell metadata including validation rules.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell (default A1)
        end_cell: Ending cell (optional, auto-expands if not provided)
        preview_only: Whether to return preview only
    
    Returns:  
    JSON string containing structured cell data with validation metadata.
    Each cell includes: address, value, row, column, and validation info (if any).
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.data import read_excel_range_with_metadata
        result = read_excel_range_with_metadata(
            full_path, 
            sheet_name, 
            start_cell, 
            end_cell
        )
        if not result or not result.get("cells"):
            return "No data found in specified range"
            
        # Return as formatted JSON string
        import json
        return json.dumps(result, indent=2, default=str)
        
    except ValueError as e:
        return f"File path error: {str(e)}"
    except Exception as e:
        logger.error(f"Error reading data: {e}")
        return f"Error reading data: {str(e)}"

@mcp.tool()
def write_data_to_excel(
    filepath: str,
    sheet_name: str,
    data: List[List],
    start_cell: str = "A1",
) -> str:
    """
    Write data to Excel worksheet.
    Excel formula will write to cell without any verification.

    PARAMETERS:  
    filepath: Path to Excel file
    sheet_name: Name of worksheet to write to
    data: List of lists containing data to write to the worksheet, sublists are assumed to be rows
    start_cell: Cell to start writing to, default is "A1"
  
    """
    try:
        full_path = get_excel_path(filepath)
        result = write_data(full_path, sheet_name, data, start_cell)
        return result["message"]
    except ValueError as e:
        return f"File path error: {str(e)}"
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error writing data: {e}")
        return f"Error writing data: {str(e)}"

@mcp.tool()
def create_workbook(filepath: str) -> str:
    """Create new Excel workbook."""
    try:
        # For create operations, allow path creation without file existence check
        if os.path.isabs(filepath):
            full_path = filepath
        else:
            if EXCEL_FILES_PATH is None:
                return f"Error: Invalid filename: {filepath}, must be an absolute path when not in SSE mode"
            full_path = os.path.join(EXCEL_FILES_PATH, filepath)
        
        # Create directory if it doesn't exist
        directory = os.path.dirname(full_path)
        if directory:
            os.makedirs(directory, exist_ok=True)
        
        from excel_mcp.workbook import create_workbook as create_workbook_impl
        create_workbook_impl(full_path)
        return f"Created workbook at {full_path}"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")
        return f"Error creating workbook: {str(e)}"

@mcp.tool()
def create_worksheet(filepath: str, sheet_name: str) -> str:
    """Create new worksheet in workbook."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_sheet as create_worksheet_impl
        result = create_worksheet_impl(full_path, sheet_name)
        return result["message"]
    except (ValidationError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating worksheet: {e}")
        return f"Error creating worksheet: {str(e)}"

@mcp.tool()
def create_chart(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = ""
) -> str:
    """Create chart in worksheet."""
    try:
        full_path = get_excel_path(filepath)
        result = create_chart_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            chart_type=chart_type,
            target_cell=target_cell,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis
        )
        return result["message"]
    except (ValidationError, ChartError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        return f"Error creating chart: {str(e)}"

@mcp.tool()
def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "mean"
) -> str:
    """Create pivot table in worksheet."""
    try:
        full_path = get_excel_path(filepath)
        result = create_pivot_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            rows=rows,
            values=values,
            columns=columns or [],
            agg_func=agg_func
        )
        return result["message"]
    except (ValidationError, PivotError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        return f"Error creating pivot table: {str(e)}"

@mcp.tool()
def create_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    table_name: Optional[str] = None,
    table_style: str = "TableStyleMedium9"
) -> str:
    """Creates a native Excel table from a specified range of data."""
    try:
        full_path = get_excel_path(filepath)
        result = create_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            table_name=table_name,
            table_style=table_style
        )
        return result["message"]
    except DataError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        return f"Error creating table: {str(e)}"

@mcp.tool()
def copy_worksheet(
    filepath: str,
    source_sheet: str,
    target_sheet: str
) -> str:
    """Copy worksheet within workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = copy_sheet(full_path, source_sheet, target_sheet)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying worksheet: {e}")
        return f"Error copying worksheet: {str(e)}"

@mcp.tool()
def delete_worksheet(
    filepath: str,
    sheet_name: str
) -> str:
    """Delete worksheet from workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = delete_sheet(full_path, sheet_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting worksheet: {e}")
        return f"Error deleting worksheet: {str(e)}"

@mcp.tool()
def rename_worksheet(
    filepath: str,
    old_name: str,
    new_name: str
) -> str:
    """Rename worksheet in workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = rename_sheet(full_path, old_name, new_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error renaming worksheet: {e}")
        return f"Error renaming worksheet: {str(e)}"

@mcp.tool()
def get_workbook_metadata(
    filepath: str,
    include_ranges: bool = False
) -> str:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        full_path = get_excel_path(filepath)
        result = get_workbook_info(full_path, include_ranges=include_ranges)
        return str(result)
    except ValueError as e:
        return f"File path error: {str(e)}"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting workbook metadata: {e}")
        return f"Error getting workbook metadata: {str(e)}"

@mcp.tool()
def merge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Merge a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        result = merge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error merging cells: {e}")
        return f"Error merging cells: {str(e)}"

@mcp.tool()
def unmerge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Unmerge a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        result = unmerge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error unmerging cells: {e}")
        return f"Error unmerging cells: {str(e)}"

@mcp.tool()
def get_merged_cells(filepath: str, sheet_name: str) -> str:
    """Get merged cells in a worksheet."""
    try:
        full_path = get_excel_path(filepath)
        return str(get_merged_ranges(full_path, sheet_name))
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting merged cells: {e}")
        return f"Error getting merged cells: {str(e)}"

@mcp.tool()
def copy_range(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None
) -> str:
    """Copy a range of cells to another location."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import copy_range_operation
        result = copy_range_operation(
            full_path,
            sheet_name,
            source_start,
            source_end,
            target_start,
            target_sheet or sheet_name  # Use source sheet if target_sheet is None
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying range: {e}")
        return f"Error copying range: {str(e)}"

@mcp.tool()
def delete_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    shift_direction: str = "up"
) -> str:
    """Delete a range of cells and shift remaining cells."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import delete_range_operation
        result = delete_range_operation(
            full_path,
            sheet_name,
            start_cell,
            end_cell,
            shift_direction
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting range: {e}")
        return f"Error deleting range: {str(e)}"

@mcp.tool()
def validate_excel_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None
) -> str:
    """Validate if a range exists and is properly formatted."""
    try:
        full_path = get_excel_path(filepath)
        range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        result = validate_range_impl(full_path, sheet_name, range_str)
        return result["message"]
    except ValidationError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating range: {e}")
        return f"Error validating range: {str(e)}"

@mcp.tool()
def get_data_validation_info(
    filepath: str,
    sheet_name: str
) -> str:
    """
    Get all data validation rules in a worksheet.
    
    This tool helps identify which cell ranges have validation rules
    and what types of validation are applied.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        
    Returns:
        JSON string containing all validation rules in the worksheet
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook
        from excel_mcp.cell_validation import get_all_validation_ranges
        
        wb = load_workbook(full_path, read_only=False)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"
            
        ws = wb[sheet_name]
        validations = get_all_validation_ranges(ws)
        wb.close()
        
        if not validations:
            return "No data validation rules found in this worksheet"
            
        import json
        return json.dumps({
            "sheet_name": sheet_name,
            "validation_rules": validations
        }, indent=2, default=str)
        
    except Exception as e:
        logger.error(f"Error getting validation info: {e}")
        return f"Error getting validation info: {str(e)}"

async def run_sse():
    """Run Excel MCP server in SSE mode."""
    # Assign value to EXCEL_FILES_PATH in SSE mode
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    # Create directory if it doesn't exist
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server with SSE transport (files directory: {EXCEL_FILES_PATH})")
        await mcp.run_sse_async()
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

def run_streamable_http():
    """Run Excel MCP server with FastAPI integration in streamable HTTP mode."""
    # Assign value to EXCEL_FILES_PATH in streamable HTTP mode
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    # Create directory if it doesn't exist
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    # Get port from environment variable, default to 8000
    port = int(os.environ.get("FASTMCP_PORT", "8000"))
    
    try:
        logger.info(f"Starting Excel MCP server with FastAPI integration (files directory: {EXCEL_FILES_PATH})")
        
        # Step 1: Create MCP ASGI app following docs pattern
        mcp_app = mcp.http_app()
        
        # Step 2: Create FastAPI app with MCP lifespan  
        app = FastAPI(
            title="Excel MCP Server API",
            description="Excel MCP Server with FastAPI integration",
            version="0.1.5",
            lifespan=mcp_app.lifespan
        )
        
        # Add CORS middleware to allow all origins and methods
        app.add_middleware(
            CORSMiddleware,
            allow_origins=["*"],  # Allow all origins
            allow_credentials=True,
            allow_methods=["*"],  # Allow all methods (GET, POST, PUT, DELETE, etc.)
            allow_headers=["*"],  # Allow all headers
        )
        
        # Step 3: Add routes directly to the FastAPI app
        @app.get("/health")
        def health_check():
            """Health check endpoint."""
            return {
                "status": "healthy",
                "service": "excel-mcp-server",
                "version": "0.1.5",
                "mcp_endpoint": "/mcp",
                "excel_files_path": EXCEL_FILES_PATH
            }
        
        @app.get("/")
        def root():
            """Root endpoint with service information."""
            return {
                "message": "Excel MCP Server API",
                "version": "0.1.5",
                "endpoints": {
                    "health": "/health",
                    "mcp": "/mcp",
                    "docs": "/docs",
                    "upload": "/upload/{user_id}",
                    "files": "/files/{user_id}"
                },
                "excel_files_path": EXCEL_FILES_PATH
            }
        
        @app.post("/upload/{user_id}")
        async def upload_excel_file(
            user_id: str,
            file: UploadFile = File(...)
        ):
            """Upload Excel file for a specific user."""
            try:
                # Validate file type
                if not file.filename:
                    raise HTTPException(status_code=400, detail="No file provided")
                
                # Check if it's an Excel file
                allowed_extensions = {'.xlsx', '.xls', '.xlsm'}
                file_extension = os.path.splitext(file.filename)[1].lower()
                if file_extension not in allowed_extensions:
                    raise HTTPException(
                        status_code=400, 
                        detail=f"Invalid file type. Allowed types: {', '.join(allowed_extensions)}"
                    )
                
                # Create user directory if it doesn't exist
                user_dir = os.path.join(EXCEL_FILES_PATH, user_id)
                os.makedirs(user_dir, exist_ok=True)
                
                # Generate file path
                file_path = os.path.join(user_dir, file.filename)
                relative_path = os.path.join(user_id, file.filename)
                
                # Save the file
                with open(file_path, "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
                
                logger.info(f"File uploaded: {file_path}")
                
                return {
                    "message": "File uploaded successfully",
                    "filename": file.filename,
                    "user_id": user_id,
                    "file_path": relative_path,  # This is what you use with MCP tools
                    "full_path": file_path,
                    "size_bytes": os.path.getsize(file_path)
                }
                
            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"Upload failed: {e}")
                raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")
        
        @app.get("/files/{user_id}")
        def list_user_files(user_id: str):
            """List all Excel files for a specific user."""
            try:
                user_dir = os.path.join(EXCEL_FILES_PATH, user_id)
                
                if not os.path.exists(user_dir):
                    return {
                        "user_id": user_id,
                        "files": [],
                        "message": "User directory does not exist"
                    }
                
                files = []
                for filename in os.listdir(user_dir):
                    file_path = os.path.join(user_dir, filename)
                    if os.path.isfile(file_path):
                        relative_path = os.path.join(user_id, filename)
                        files.append({
                            "filename": filename,
                            "relative_path": relative_path,  # Use this with MCP tools
                            "size_bytes": os.path.getsize(file_path),
                            "modified": os.path.getmtime(file_path)
                        })
                
                return {
                    "user_id": user_id,
                    "files": files,
                    "count": len(files)
                }
                
            except Exception as e:
                logger.error(f"Failed to list files for user {user_id}: {e}")
                raise HTTPException(status_code=500, detail=f"Failed to list files: {str(e)}")
        
        @app.delete("/files/{user_id}/{filename}")
        def delete_user_file(user_id: str, filename: str):
            """Delete a specific file for a user."""
            try:
                file_path = os.path.join(EXCEL_FILES_PATH, user_id, filename)
                
                if not os.path.exists(file_path):
                    raise HTTPException(status_code=404, detail="File not found")
                
                os.remove(file_path)
                logger.info(f"File deleted: {file_path}")
                
                return {
                    "message": "File deleted successfully",
                    "filename": filename,
                    "user_id": user_id
                }
                
            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"Failed to delete file: {e}")
                raise HTTPException(status_code=500, detail=f"Failed to delete file: {str(e)}")
        
        # Step 4: Mount the MCP server at root path to handle /mcp/mcp/ requests
        app.mount("/mcp", mcp_app)
        
        logger.info(f"Server will be available at:")
        logger.info(f"  - API Health Check: http://localhost:{port}/health")
        logger.info(f"  - API Root: http://localhost:{port}/")
        logger.info(f"  - MCP Endpoint: http://localhost:{port}/mcp")
        logger.info(f"  - API Docs: http://localhost:{port}/docs")
        
        # Step 5: Run with uvicorn
        uvicorn.run(app, host="0.0.0.0", port=port, log_level="info")
        
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

def run_stdio():
    """Run Excel MCP server in stdio mode."""
    # No need to assign EXCEL_FILES_PATH in stdio mode
    
    try:
        logger.info("Starting Excel MCP server with stdio transport")
        mcp.run(transport="stdio")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")